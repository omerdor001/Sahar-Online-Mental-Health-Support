import collections
import logging
import os
import time
import threading
from datetime import datetime, timedelta, date
import json
from zipfile import BadZipFile
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from time import sleep
from AI.algorithms.summerization_algorithm import Summarizer
from DataBase.database_helper import DataBaseHelper
from dao_object.conversation_prediction_object import ConversationPredictionDAO
from utils.config_util import ConfigUtil
from objects.conversation_history_record import ConsumerParticipant, ConversationHistoryRecord, ConversationSummary
from objects.message_record import MessageRecord
from lp_api_manager.lp_utils import LpUtils
from readerwriterlock import rwlock


class ConversationCache:
    _instance = None
    closed_conversations: dict[str, ConversationHistoryRecord]
    messages: dict[str, set[MessageRecord]]
    open_conversations: dict[str, ConversationHistoryRecord]
    missing_conversations: dict[str, int]  # Track how many times a conversation has been missing
    converstaion_summaries: dict[str, ConversationSummary]

        
    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls, *args, **kwargs)
            cls._instance.closed_conversations = {}
            cls._instance.messages = {}
            cls._instance.open_conversations = {}
            cls._instance.missing_conversations = {}  # Initialize missing conversations counter
            cls._instance.converstaion_summaries = {}
            cls._instance._lock = rwlock.RWLockFair()
            cls._instance._start_cleanup_thread()
            cls._instance._start_log_thread()
        return cls._instance
    
    def initialize(self, initial_predictions=None): # Nir Add it
        if initial_predictions:
            with self._lock.gen_wlock():



                    def chunk_list(lst, chunk_size):
                                for i in range(0, len(lst), chunk_size):
                                    yield lst[i:i + chunk_size]
                            
                    for chunk in chunk_list(initial_predictions, 100):
                        conversation_ids = [prediction.conversationId for prediction in chunk]    #check
                        response = LpUtils().get_conversations_by_conv_id(conversation_ids)
                        conversations_records: dict[str, ConversationHistoryRecord] = (
                            LpUtils.extract_conversations(response)
                        )

                        messages_records = LpUtils.extract_messages(response)

                        for prediction in chunk:
                            conv_id = prediction.conversationId
                            conv_record = conversations_records[conv_id]
                            conv_record.GSR = prediction.GSR
                            conv_record.basic_algorithm_result = prediction.basic_algorithm_result
                            conv_record.IMSR = prediction.IMSR
                            conv_record.max_GSR = prediction.max_GSR
                            if conv_record.status == "CLOSE" or conv_record.status == "closed" or conv_record.status == "CLOSED" or conv_record.status == "Closed":
                                self.closed_conversations[conv_id] = conv_record
                            else:
                                self.open_conversations[conv_id] = conv_record


                        for conv_id, message_list in messages_records.items():
                            self.messages[conv_id] = message_list



    def _start_cleanup_thread(self):
        self.cleanup_thread = threading.Thread(target=self.cleanup_old_data)
        self.cleanup_thread.daemon = True
        self.cleanup_thread.start()

    def _start_log_thread(self):
        self.cleanup_thread = threading.Thread(target=self.log_data)
        self.cleanup_thread.daemon = True
        self.cleanup_thread.start()

    def update_conversations(
            self,
            currentTime,
            conversation: dict[str, ConversationHistoryRecord],
            messages: dict[str, set[MessageRecord]],
    ):
        logging.info("Start cache update")
        move_to_close: [str] = []
        with self._lock.gen_wlock():
            # Reset missing counter for conversations that are present
            for conv_id in conversation.keys():
                if conv_id in self.missing_conversations:
                    del self.missing_conversations[conv_id]

            # Check for missing conversations
            for conv_id, conv_record in self.open_conversations.items():
                if conv_id not in conversation.keys() and conv_record.brandId != 'testing':
                    # Increment missing counter
                    self.missing_conversations[conv_id] = self.missing_conversations.get(conv_id, 0) + 1
                    
                    # Only move to close if missing for 3 consecutive times
                    if self.missing_conversations[conv_id] >= 5:
                        logging.info(f"conversation {conv_id} has been missing for {self.missing_conversations[conv_id]} times, moving to closed")
                        move_to_close.append(conv_id)
                        del self.missing_conversations[conv_id]
                    else:
                        logging.info(f"conversation {conv_id} is temporarily missing (count: {self.missing_conversations[conv_id]})")

            self._move_conv_to_close(currentTime, move_to_close)

            # updating open conversations dict
            for conv_id, conv_record in conversation.items():
                new_record = self.open_conversations.get(conv_id)
                if new_record is None:
                    new_record = conv_record
                else:
                    last_message_id = max(messages[conv_id], key=lambda msg: msg.timeL if msg.timeL else 0).messageId
                    new_record.update(conv_record, last_message_id)
                new_record.update_field("lastUpdateTime", currentTime)
                self.open_conversations[conv_id] = new_record

            for conv_id, message_list in messages.items():
                self.messages[conv_id] = message_list
        logging.info("Finish cache update Successfully")

    def update_conversations_test(
            self,
            currentTime,
            conversation: dict[str, ConversationHistoryRecord],
            messages: dict[str, set[MessageRecord]],
    ):
        logging.info("Start cache update")
        move_to_close: [str] = []
        with self._lock.gen_wlock():
            # moving all new conversation that closed

            # updating open conversations dict
            for conv_id, conv_record in conversation.items():
                new_record = self.open_conversations.get(conv_id)
                if new_record is None:
                    new_record = conv_record
                else:
                    new_record.update(conv_record)
                new_record.update_field("need_analyze", True)    
                new_record.update_field("lastUpdateTime", currentTime)
                self.open_conversations[conv_id] = new_record

            for conv_id, message_list in messages.items():
                self.messages[conv_id] = message_list
        logging.info("Finish cache update Successfully")
    
    def update_conversations_test_closed(
            self,
            currentTime,
            conversation: dict[str, ConversationHistoryRecord],
            messages: dict[str, set[MessageRecord]],
    ):
        logging.info("Start cache update")
        move_to_close: [str] = []
        with self._lock.gen_wlock():
            # moving all new conversation that closed

            # updating open conversations dict
            for conv_id, conv_record in conversation.items():
                conv_record.update_field("need_analyze", True)
                self.closed_conversations[conv_id] = conv_record
            
            for conv_id, message_list in messages.items():
                self.messages[conv_id] = message_list
        logging.info("Finish cache update Successfully")

    def _move_conv_to_close(self, now, conversations_ids: [str]):
        if len(conversations_ids) == 0:
            return

        response = LpUtils().get_conversations_by_conv_id(conversations_ids)
        conversations_records: dict[str, ConversationHistoryRecord] = (
            LpUtils.extract_conversations(response)
        )
        messages_records = LpUtils.extract_messages(response)

        for conv_id, conv_record in conversations_records.items():
            current_conversation: ConversationHistoryRecord = self.open_conversations.get(conv_id)
            current_conversation.update(conv_record)
            current_conversation.update_field("need_analyze", True)
            current_conversation.update_field("lastUpdateTime", now)
            self.closed_conversations[conv_id] = current_conversation
            self.open_conversations.pop(conv_id, None)

        for conv_id, message_list in messages_records.items():
            self.messages[conv_id] = message_list

    def get_all_conversations(self) -> set[ConversationHistoryRecord]:
        all_info = set()
        with self._lock.gen_rlock():
            for conv_id, conv_data in self.open_conversations.items():
                all_info.add(conv_data)
            for conv_id, conv_data in self.closed_conversations.items():
                all_info.add(conv_data)
        return all_info

    def get_closed_conversations(self) -> dict[str, ConversationHistoryRecord]:
        with self._lock.gen_rlock():
            return self.closed_conversations

    def get_conversation_by_Id(self, conv_id):
        with self._lock.gen_rlock():
            if conv_id in self.open_conversations:
                conversation = self.open_conversations.get(conv_id)
            else:
                conversation = self.closed_conversations.get(conv_id)
        if conversation:
            return conversation
        else:
            logging.warning(f"conversation {conv_id} not found")
            return None

    def get_conversations_to_enrich(self) -> dict[str, [MessageRecord]]:
        with self._lock.gen_rlock():
            conversations: dict[str, [MessageRecord]] = {}

            for conv_id, record in self.open_conversations.items():
                if record.need_analyze:
                    conversations[conv_id] = self.messages.get(conv_id)

            for conv_id, record in self.closed_conversations.items():
                if record.need_analyze:
                    conversations[conv_id] = self.messages.get(conv_id)


        return conversations

    def get_open_conversations(self) -> dict[str, ConversationHistoryRecord]:
        with self._lock.gen_rlock():
            return self.open_conversations

    def extract_hebrew_messages(self, messages: list[MessageRecord]) -> list[MessageRecord]:
        return [msg for msg in messages if msg.messageData['msg']['text']]

    def extract_hebrew_test(
        self, conversation : ConversationHistoryRecord
    ) -> str:
        messages = self.messages.get(conversation.conversationId)
        lines = []
        if not messages:
            return ""
        messages = sorted(messages, key=lambda m: (m.timeL is None, m.timeL))
        for msg in messages:
            if hasattr(msg, "messageId") and msg.messageId:
                msg_id_str = msg.messageId.split(":")[-1]
        for msg in messages:
            # Check for RICH_CONTENT type
            if getattr(msg, "type", None) == "RICH_CONTENT":
                rich_content = msg.messageData.get("richContent") if msg.messageData else None
                if rich_content and "content" in rich_content:
                    content = rich_content["content"]
                    if isinstance(content, str) and "גילך" in content:
                        lines.append("מה טווח גילך?")
                # Ignore all other RICH_CONTENT messages
                continue
            # Regular messages with text
            text = None
            if msg.messageData and "msg" in msg.messageData:
                msg_obj = msg.messageData["msg"]
                if isinstance(msg_obj, dict):
                    text = msg_obj.get("text")
            if text:
                prefix = "נציג:" if msg.sentBy == "Agent" else "פונה:"
                lines.append(f"{prefix} {text}")
        text = "\n".join(lines)
        return text


    def get_summary(self, conversation_id: str):
        
        summary = None
        last_message_id = ""
        with self._lock.gen_rlock():
            conversation = self.get_conversation_by_Id(conversation_id)
            last_message_id = conversation.last_effected_message_id
            if conversation:
                if conversation_id in self.converstaion_summaries and conversation.last_effected_message_id == self.converstaion_summaries[conversation_id].last_message_id:
                    return self.converstaion_summaries[conversation_id]
                messages = self.messages.get(conversation_id)
                if messages:
                    text = self.extract_hebrew_test(conversation)
                    if text != "":
                        summarizer = Summarizer()
                        summary = summarizer.get_summary(text)
        with self._lock.gen_wlock():
            if summary:
                self.converstaion_summaries[conversation_id] = ConversationSummary(last_message_id, summary)
        if summary:
            return self.converstaion_summaries[conversation_id] 


    def get_json_open_conversations(self) -> list[ConversationHistoryRecord]:
        json_open_conversations = ConversationCache().get_open_conversations()
        for c in json_open_conversations.values():
            logging.info(f"conversation {c.conversationId} inside get_json_open_conversations")
        json_open_conversations = [
            conv.to_dict() for conv in json_open_conversations.values()
        ]
        return json_open_conversations

    def update_need_analyze(self, data):
        with self._lock.gen_wlock():
            for conv_id in data:
                conv_record = self.open_conversations.get(conv_id)
                if conv_record:
                    conv_record.update_field("need_analyze", False)
    
    def enrich_conversation(self, conv_id: str, json_str: str, last_message_id: str):
        logging.debug(f"Start enriching conversation: {conv_id}")
        with self._lock.gen_wlock():
            try:

                from_open = False
                # Get the existing conversation record
                conv_record = self.closed_conversations.get(conv_id)
                if conv_record is None:
                    conv_record = self.open_conversations.get(conv_id)
                    from_open = True

                if conv_record is None:
                    logging.warning(f"Conversation {conv_id} not found.")
                    return

                try:
                    data = json.loads(json_str)
                except json.JSONDecodeError as e:
                    logging.error(f"JSON decode error for {conv_id}: {e}")
                    return
                # Add last_effected_message_id to json_data
                data['last_effected_message_id'] = last_message_id


                push_to_db = False
                if from_open:
                    
                    if self.open_conversations[conv_id].last_effected_message_id != last_message_id:
                        push_to_db = True
                else:
                    push_to_db = True


                # Update the conversation record with the fields from the JSON
                for key, value in data.items():
                    if key == "GSR":
                        logging.info(f"Updating GSR for conversation {conv_id} to {value}")
                        logging.info(f"type of value is {type(value)}")
                    conv_record.update_field(key, value)

                # Update the cache with the modified conversation record
                if from_open:
                    logging.info(f"enrich conversation {conv_id} open")
                    self.open_conversations[conv_id] = conv_record
                else:
                    logging.info(f"enrich conversation {conv_id} closed")
                    conv_record.update_field("need_analyze", False)
                    self.closed_conversations[conv_id] = conv_record

                if push_to_db:
                    DataBaseHelper.add_b(ConversationPredictionDAO(conversationId=conv_record.conversationId,
                                                   basic_algorithm_result=conv_record.basic_algorithm_result,
                                                    GSR=conv_record.GSR,IMSR=conv_record.IMSR,
                                                    max_GSR=conv_record.max_GSR, status=conv_record.status,last_message_id=conv_record.last_effected_message_id))

            except Exception as e:
                logging.error(f"Unexpected error enriching conversation {conv_id}: {e}")
        
        logging.debug(f"End enriching conversation: {conv_id}, successfully")

    def log_data(self):
        while True:
            try:
                time.sleep(
                    int(ConfigUtil().get_config_attribute("logCacheInterval")))  # Sleep for the configured interval
                logging.debug("Start log cache data")

                file_path = f"logs/cache_log_{date.today().strftime('%m_%d_%y')}.xlsx"
                with self._lock.gen_rlock():
                    self.log_summary_line(file_path)
                    self.log_conversation_info(file_path)
                logging.debug("Finish log cache data successfully")

            except Exception as e:
                logging.error(f"Error in logging cache data: {e}")


    def log_summary_line(self, file_path):
        current_time = datetime.now()
        count_open_calls = len(self.open_conversations)
        count_closed_calls = len(self.closed_conversations)

        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb = Workbook()
        except (InvalidFileException, BadZipFile) as e:
            logging.error(f"Invalid file: {file_path}. {e}")
            wb = self.DR_cache_file(file_path)

        if wb is None:
            return

        sheet = wb.active

        if sheet.max_row == 1:
            row = 1
        else:
            row = sheet.max_row + 2  # Start appending from the next row after existing content

        sheet.cell(row, 1, f"Time: {current_time}")
        sheet.cell(row, 2, f"Number of current Open Calls: {count_open_calls}")
        sheet.cell(row, 3, f"Number of current Closed Calls: {count_closed_calls}")

        wb.save(file_path)

    def log_conversation_info(self, file_path):
        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            wb = Workbook()
        except (InvalidFileException, BadZipFile) as e:
            logging.error(f"Invalid file: {file_path}. {e}")
            wb = self.DR_cache_file(file_path)

        if wb is None:
            return

        sheet = wb.active

        # Write header row if the sheet is newly created
        header = [
            "conversation ID",
            "status",
            "start time",
            "end time",
            "agent ID",
            "agent full name",
            "GSR Algorithm",
            "Max GSR",
            "last_effected_message_id"
        ]
        row = sheet.max_row + 1
        for col, header_value in enumerate(header, start=1):
            sheet.cell(row, col, header_value)

        # Find the next available row to append data
        row = row + 1

        # Write data for open conversations
        for conv in self.open_conversations.values():
            sheet.cell(row, 1, conv.conversationId)
            sheet.cell(row, 2, conv.status)
            sheet.cell(row, 3, conv.startTime)
            sheet.cell(row, 4, conv.conversationEndTime)
            sheet.cell(row, 5, conv.latestAgentId)
            sheet.cell(row, 6, conv.latestAgentFullName)
            sheet.cell(row, 7, conv.GSR)
            sheet.cell(row, 8, conv.max_GSR)
            sheet.cell(row, 9, conv.last_effected_message_id)
            row += 1

        # Write data for closed conversations
        for conv in self.closed_conversations.values():
            sheet.cell(row, 1, conv.conversationId)
            sheet.cell(row, 2, conv.status)
            sheet.cell(row, 3, conv.startTime)
            sheet.cell(row, 4, conv.conversationEndTime)
            sheet.cell(row, 5, conv.latestAgentId)
            sheet.cell(row, 6, conv.latestAgentFullName)
            sheet.cell(row, 7, conv.GSR)
            sheet.cell(row, 8, conv.max_GSR)
            sheet.cell(row, 9, conv.last_effected_message_id)
            row += 1

        wb.save(file_path)

    def DR_cache_file(self, file_path):
        # Generate a new file name with a timestamp
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        new_file_path = f"{file_path}_{timestamp}.bak"

        try:
            # Rename the existing corrupted file
            os.rename(file_path, new_file_path)
            logging.info(f"Renamed corrupted file to: {new_file_path}")
        except FileNotFoundError:
            logging.warning(f"File not found: {file_path}. No file to rename.")
        except Exception as e:
            logging.error(f"Error renaming file: {e}")

        try:
            # Create a new, valid Excel file
            wb = Workbook()
            wb.save(file_path)
            logging.info(f"Created a new file: {file_path}")
            return wb
        except Exception as e:
            logging.error(f"Error creating a new file: {e}")



    def cleanup_old_data(self):
        while True:
            try:
                time.sleep(
                    ConfigUtil().get_config_attribute("deleteThreadCacheInterval")
                )  # Sleep for 2 hours before performing the cleanup again
                current_time = datetime.now()
                logging.info("Start clean old cache data")
                with self._lock.gen_wlock():
                    copy_closed_conversations = self.closed_conversations.items()
                    for conv_id, conv_data in copy_closed_conversations:
                        if isinstance(conv_data, ConversationHistoryRecord):
                            last_update_time = datetime.fromtimestamp(
                                conv_data.lastUpdateTime / 1000
                            )
                            max_cache_age = ConfigUtil().get_config_attribute("cacheDataMaxAge")
                            if current_time - last_update_time > timedelta(
                                    seconds=max_cache_age
                            ):
                                logging.debug(f"Remove old conversation: {conv_id}")
                                self.closed_conversations.pop(conv_id)
                                self.messages.pop(conv_id)
                    logging.info("finish clean old cache data successfully")
            except ValueError as e:
                logging.error(f"Cant read the value deleteThreadCacheInterval from config. {e}")
                time.sleep(7200)
            except Exception as e:
                logging.error(f"error in clean old cache data. {e}")

